# app_gui.py
import sys, io, datetime, os, traceback
import pandas as pd

from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex, QThread, Signal
from PySide6.QtGui import QAction, QIcon
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
    QLabel, QPushButton, QVBoxLayout, QHBoxLayout, QGroupBox, QProgressBar,
    QLineEdit, QDateEdit, QTabWidget, QComboBox, QInputDialog, QTableView
)

# Thêm sau các import sẵn có:
from rapidfuzz import process, fuzz
from unidecode import unidecode
import re

def resource_path(rel_path: str) -> str:
    # Khi chạy bằng PyInstaller one-file, tài nguyên nằm trong _MEIPASS
    base = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base, rel_path)
# ======= THEME (đổi màu ở đây) =======
PRIMARY = "#987049"   # màu thương hiệu bạn đang dùng trong bản streamlit
PRIMARY_DARK = "#7a593a"
BG = "#f7f6f4"
CARD_BG = "#ffffff"
TXT = "#222222"

APP_STYLES = f"""
    QWidget {{
        background: {BG};
        color: {TXT};
        font-size: 13px;
    }}
    QGroupBox {{
        background: {CARD_BG};
        border: 1px solid rgba(0,0,0,0.08);
        border-radius: 12px;
        margin-top: 12px;
        padding: 12px;
    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        left: 10px;
        padding: 0 6px;
        color: {PRIMARY};
        font-weight: 600;
        background: {BG};
        border-radius: 6px;
    }}
    QLabel {{
        font-weight: 500;
    }}
    QLineEdit {{
        background: #fff;
        border: 1px solid rgba(0,0,0,0.15);
        border-radius: 10px;
        padding: 8px 10px;
    }}
    QLineEdit:focus {{
        border: 1px solid {PRIMARY};
        outline: none;
    }}
    QDateEdit {{
        background: #fff;
        border: 1px solid rgba(0,0,0,0.15);
        border-radius: 10px;
        padding: 6px 8px;
    }}
    QPushButton {{
        background: {PRIMARY};
        color: white;
        border: none;
        padding: 10px 14px;
        border-radius: 12px;
        font-weight: 600;
    }}
    QPushButton:hover {{
        background: {PRIMARY_DARK};
    }}
    QPushButton:disabled {{
        background: #c7b7a7;
    }}
    QProgressBar {{
        border: 1px solid rgba(0,0,0,0.1);
        border-radius: 10px;
        background: #f0ebe6;
        text-align: center;
        height: 18px;
    }}
    QProgressBar::chunk {{
        background-color: {PRIMARY};
        border-radius: 10px;
    }}
    QTabWidget::pane {{
        border: 1px solid rgba(0,0,0,0.08);
        border-radius: 12px;
        padding: 6px;
        background: {CARD_BG};
    }}
    QTabBar::tab {{
        background: #efe7df;
        border: 1px solid rgba(0,0,0,0.08);
        padding: 8px 14px;
        border-top-left-radius: 10px;
        border-top-right-radius: 10px;
        margin-right: 6px;
    }}
    QTabBar::tab:selected {{
        background: {PRIMARY};
        color: white;
        font-weight: 600;
    }}
"""

# ========== Utilities (giữ nguyên nghiệp vụ) ==========
"""def parse_vietnamese_date(value: str) -> pd.Timestamp:
    if isinstance(value, str):
        parts = value.strip().split()
        if len(parts) == 4 and parts[1].lower() == 'thg':
            day, _, month, year = parts
            date_str = f"{day}/{month}/{year}"
            return pd.to_datetime(date_str, format="%d/%m/%Y", errors="coerce")
        return pd.to_datetime(value, dayfirst=True, errors='coerce')
    return pd.NaT

def detect_header_row(df_raw):
    for idx, row in df_raw.iterrows():
        if row.astype(str).str.contains('STT', na=False).any():
            return idx
    raise ValueError("Không tìm thấy dòng header chứa 'STT'")

def load_and_flatten_eas(eas_bytes):
    df_raw = pd.read_excel(io.BytesIO(eas_bytes), header=None)
    df_raw.iloc[:, 0] = df_raw.iloc[:, 0].astype(str)
    df_raw = df_raw[~df_raw.iloc[:, 0].str.contains(r'^\[\d+\]$', na=False)].reset_index(drop=True)
    header_row = detect_header_row(df_raw)
    df = pd.read_excel(io.BytesIO(eas_bytes), header=[header_row, header_row+1])

    flat_cols = []
    for top, sub in df.columns:
        if pd.notna(sub) and not str(sub).startswith("Unnamed"):
            flat_cols.append(str(sub).strip())
        else:
            flat_cols.append(str(top).strip())
    df.columns = flat_cols
    return df

def clean_eas(df):
    rename_map = {
        'Tên người mua(Buyer Name)': 'Buyer Name',
        'Ngày, tháng, năm phát hành': 'ISSUE_DATE',
        'Doanh số bán chưa có thuế(Revenue excluding VAT)': 'Revenue_ex_VAT',
        'Thuế GTGT(VAT amount)': 'VAT_Amount',
        'Ký hiệu mẫu hóa đơn': 'InvoiceSerial',
        'Số hóa đơn': 'InvoiceNumber'
    }
    df = df.rename(columns=rename_map)
    mst_col = next((c for c in df.columns if 'Mã số thuế' in c or 'Tax code' in c), None)
    if mst_col:
        df = df.rename(columns={mst_col: 'TaxCode'})
    df = df.dropna(subset=['Buyer Name', 'Revenue_ex_VAT']).reset_index(drop=True)
    return df

def build_fiv(df_eas, df_kh):
    taxkey_kh = next((c for c in df_kh.columns if any(x in c for x in ['MST','CMND','PASSPORT','Tax code'])), None)
    records = []
    for idx, row in df_eas.iterrows():
        buyer = row['Buyer Name']
        cust_acc = pd.NA
        if 'TaxCode' in row and pd.notna(row['TaxCode']) and taxkey_kh:
            m = df_kh[df_kh[taxkey_kh] == row['TaxCode']]['Customer account']
            if not m.empty:
                cust_acc = m.iat[0]
        if pd.isna(cust_acc):
            m = df_kh[df_kh['Name'] == buyer]['Customer account']
            if not m.empty:
                cust_acc = m.iat[0]
        line_amount = row['Revenue_ex_VAT']
        vat_amount  = row.get('VAT_Amount', 0)
        total_amt   = line_amount + vat_amount
        records.append({
            'IdRef': idx + 1,
            'InvoiceDate': row['ISSUE_DATE'],
            'DocumentDate': row['ISSUE_DATE'],
            'CurrencyCode': 'VND',
            'CustAccount': cust_acc,
            'InvoiceAccount': cust_acc,
            'SalesName': buyer,
            'APMA_DimA': 'TX',
            'APMC_DimC': '0000',
            'APMD_DimD': '00',
            'APMF_DimF': '0000',
            'TaxGroupHeader': 'OU',
            'PostingProfile': '131103',
            'LineNum': 1,
            'Description': 'Doanh thu dịch vụ spa',
            'SalesPrice': line_amount,
            'SalesQty': 1,
            'LineAmount': line_amount,
            'TaxAmount': vat_amount,
            'TotalAmount': total_amt,
            'TaxGroupLine': 'OU',
            'TaxItemGroup': '10%',
            'Line_MainAccountId': '511301',
            'Line_APMA_DimA': 'TX',
            'Line_APMC_DimC': '5301',
            'Line_APMD_DimD': '00',
            'Line_APMF_DimF': '0000',
            'BHS_VATInvocieDate_VATInvoice': row['ISSUE_DATE'],
            'BHS_Form_VATInvoice': '',
            'BHS_Serial_VATInvoice': row.get('InvoiceSerial', ''),
            'BHS_Number_VATInvoice': row.get('InvoiceNumber', ''),
            'BHS_Description_VATInvoice': 'Doanh thu dịch vụ spa'
        })

    cols_order = [
        'IdRef','InvoiceDate','DocumentDate','CurrencyCode','CustAccount','InvoiceAccount',
        'SalesName','APMA_DimA','APMC_DimC','APMD_DimD','APMF_DimF','TaxGroupHeader',
        'PostingProfile','LineNum','Description','SalesPrice','SalesQty','LineAmount',
        'TaxAmount','TotalAmount','TaxGroupLine','TaxItemGroup','Line_MainAccountId',
        'Line_APMA_DimA','Line_APMC_DimC','Line_APMD_DimD','Line_APMF_DimF',
        'BHS_VATInvocieDate_VATInvoice','BHS_Form_VATInvoice','BHS_Serial_VATInvoice',
        'BHS_Number_VATInvoice','BHS_Description_VATInvoice'
    ]
    return pd.DataFrame(records, columns=cols_order)"""
# ========== Utilities (giữ nguyên nghiệp vụ) ==========

def parse_vietnamese_date(value: str) -> pd.Timestamp:
    if isinstance(value, str):
        parts = value.strip().split()
        if len(parts) == 4 and parts[1].lower() == 'thg':
            day, _, month, year = parts
            date_str = f"{day}/{month}/{year}"
            return pd.to_datetime(date_str, format="%d/%m/%Y", errors="coerce")
        return pd.to_datetime(value, dayfirst=True, errors='coerce')
    return pd.NaT

def detect_header_row(df_raw):
    for idx, row in df_raw.iterrows():
        if row.astype(str).str.contains('STT', na=False).any():
            return idx
    raise ValueError("Không tìm thấy dòng header chứa 'STT'")

def load_and_flatten_eas(eas_bytes):
    df_raw = pd.read_excel(io.BytesIO(eas_bytes), header=None)
    df_raw.iloc[:, 0] = df_raw.iloc[:, 0].astype(str)
    df_raw = df_raw[~df_raw.iloc[:, 0].str.contains(r'^\[\d+\]$', na=False)].reset_index(drop=True)
    header_row = detect_header_row(df_raw)
    df = pd.read_excel(io.BytesIO(eas_bytes), header=[header_row, header_row+1])

    flat_cols = []
    for top, sub in df.columns:
        if pd.notna(sub) and not str(sub).startswith("Unnamed"):
            flat_cols.append(str(sub).strip())
        else:
            flat_cols.append(str(top).strip())
    df.columns = flat_cols
    return df

def clean_eas(df):
    rename_map = {
        'Tên người mua(Buyer Name)': 'Buyer Name',
        'Ngày, tháng, năm phát hành': 'ISSUE_DATE',
        'Doanh số bán chưa có thuế(Revenue excluding VAT)': 'Revenue_ex_VAT',
        'Thuế GTGT(VAT amount)': 'VAT_Amount',
        'Ký hiệu mẫu hóa đơn': 'InvoiceSerial',
        'Số hóa đơn': 'InvoiceNumber'
    }
    df = df.rename(columns=rename_map)
    mst_col = next((c for c in df.columns if 'Mã số thuế' in c or 'Tax code' in c), None)
    if mst_col:
        df = df.rename(columns={mst_col: 'TaxCode'})
    df = df.dropna(subset=['Buyer Name', 'Revenue_ex_VAT']).reset_index(drop=True)
    return df

# ==== FIX: luôn normalize account code về text có leading zero ====
def normalize_account_code(val, width=9):
    if pd.isna(val):
        return pd.NA
    s = str(val).strip()
    if re.match(r'^\d+(\.0+)?$', s):
        s = s.split('.')[0]
    s = re.sub(r'\D', '', s)
    return s.zfill(width)

def build_fiv(df_eas, df_kh):
    # chuẩn hóa cột Customer account trong file KH
    if 'Customer account' in df_kh.columns:
        df_kh['Customer account'] = df_kh['Customer account'].apply(lambda x: normalize_account_code(x, 9))

    taxkey_kh = next((c for c in df_kh.columns if any(x in c for x in ['MST','CMND','PASSPORT','Tax code'])), None)
    records = []
    for idx, row in df_eas.iterrows():
        buyer = row['Buyer Name']
        cust_acc = pd.NA

        if 'TaxCode' in row and pd.notna(row['TaxCode']) and taxkey_kh:
            m = df_kh[df_kh[taxkey_kh] == row['TaxCode']]['Customer account']
            if not m.empty:
                cust_acc = m.iat[0]

        if pd.isna(cust_acc):
            m = df_kh[df_kh['Name'] == buyer]['Customer account']
            if not m.empty:
                cust_acc = m.iat[0]

        # luôn normalize cust_acc
        if pd.notna(cust_acc):
            cust_acc = normalize_account_code(cust_acc, 9)

        line_amount = row['Revenue_ex_VAT']
        vat_amount  = row.get('VAT_Amount', 0)
        total_amt   = line_amount + vat_amount

        # ==== xử lý InvoiceNumber để loại .0 ====
        raw_inv = row.get('InvoiceNumber', '')
        if pd.notna(raw_inv) and raw_inv != '':
            try:
                inv_number = str(int(float(raw_inv)))  # 145.0 -> 145
            except Exception:
                inv_number = str(raw_inv).strip()
        else:
            inv_number = ''

        records.append({
            'IdRef': idx + 1,
            'InvoiceDate': row['ISSUE_DATE'],
            'DocumentDate': row['ISSUE_DATE'],
            'CurrencyCode': 'VND',
            'CustAccount': cust_acc,
            'InvoiceAccount': cust_acc,
            'SalesName': buyer,
            'APMA_DimA': 'TX',
            'APMC_DimC': '0000',
            'APMD_DimD': '00',
            'APMF_DimF': '0000',
            'TaxGroupHeader': 'OU',
            'PostingProfile': '131103',
            'LineNum': 1,
            'Description': 'Doanh thu dịch vụ spa',
            'SalesPrice': line_amount,
            'SalesQty': 1,
            'LineAmount': line_amount,
            'TaxAmount': vat_amount,
            'TotalAmount': total_amt,
            'TaxGroupLine': 'OU',
            'TaxItemGroup': '10%',
            'Line_MainAccountId': '511301',
            'Line_APMA_DimA': 'TX',
            'Line_APMC_DimC': '5301',
            'Line_APMD_DimD': '00',
            'Line_APMF_DimF': '0000',
            'BHS_VATInvocieDate_VATInvoice': row['ISSUE_DATE'],
            'BHS_Form_VATInvoice': '',
            'BHS_Serial_VATInvoice': row.get('InvoiceSerial', ''),
            'BHS_Number_VATInvoice': inv_number,
            'BHS_Description_VATInvoice': 'Doanh thu dịch vụ spa'
        })

    cols_order = [
        'IdRef','InvoiceDate','DocumentDate','CurrencyCode','CustAccount','InvoiceAccount',
        'SalesName','APMA_DimA','APMC_DimC','APMD_DimD','APMF_DimF','TaxGroupHeader',
        'PostingProfile','LineNum','Description','SalesPrice','SalesQty','LineAmount',
        'TaxAmount','TotalAmount','TaxGroupLine','TaxItemGroup','Line_MainAccountId',
        'Line_APMA_DimA','Line_APMC_DimC','Line_APMD_DimD','Line_APMF_DimF',
        'BHS_VATInvocieDate_VATInvoice','BHS_Form_VATInvoice','BHS_Serial_VATInvoice',
        'BHS_Number_VATInvoice','BHS_Description_VATInvoice'
    ]
    out = pd.DataFrame(records, columns=cols_order)

    # ép các cột về string để giữ nguyên khi export
    for c in ['CustAccount', 'InvoiceAccount', 'BHS_Number_VATInvoice']:
        out[c] = out[c].astype('string')

    return out



# ==============================
# Helpers cho Hospital Price Consolidator
# ==============================
def normalize_text(s: str) -> str:
    if pd.isna(s): return ""
    s = str(s).strip().lower()
    s = unidecode(s)
    s = re.sub(r"[ \t\r\n]+", " ", s)
    s = re.sub(r"[\-_,.;:()\[\]{}]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

CANDIDATE_COLS = {
    "service": ["dich vu","ten dich vu","noi dung","chi dinh","dich vu kham","ten chi dinh", "ten xet nghiem","hang muc","hang muc kham","noi dung kham","kham","dv","ten dv","ten"],
    "price":   ["don gia","gia","gia tien","bao gia","unit price","price","gia dich vu"],
    "unit":    ["don vi","unit","dv","don vi tinh"],
    "dept":    ["khoa","chuyen khoa","phong","bo phan"],
    "note":    ["ghi chu","note","mo ta","dien giai"],
    "vat":     ["vat","thue","thue gtgt"]
}
def guess_colname(col: str):
    c = normalize_text(col)
    for target, keys in CANDIDATE_COLS.items():
        for k in keys:
            if k in c: return target
    return None

DEFAULT_SYNONYMS = {
    "kham tong quat": ["kham suc khoe tong quat","goi kham tong quat","kham tong","kham sk"],
    "xet nghiem mau": ["xet nghiem huyet hoc","xn mau","xet nghiem tong quat mau","xet nghiem sinh hoa mau"],
    "sieu am bung": ["sieu am bung tong quat","sieu am o bung","sieu am o bung tong quat"],
    "x quang": ["xquang","x - quang","chup xquang"],
    "mri": ["cong huong tu","chup mri"],
    "ct": ["cat lop vi tinh","chup ct","ct scanner"],
    "noi soi da day": ["noi soi da day khong gay me","noi soi da day co gay me"],
    "noi soi dai trang": ["noi soi dai trang khong gay me","noi soi dai trang co gay me"],
    "sieu am tim": ["echo tim","sieu am doppler tim"],
    "dien tim": ["ecg","dientim","dtd"],
    "sieu am tuyen giap": ["sieu am tuyen giap doppler"],
    "kham nhi": ["kham tre em","kham nhi khoa"],
    "kham tai mui hong": ["kham tmh","tai mui hong"],
    "kham mat": ["kham nhan khoa","kham mat tong quat"],
}
def apply_synonyms(s: str, synonyms: dict) -> str:
    base = normalize_text(s)
    for canon, alts in synonyms.items():
        if base == canon: return canon
        for a in alts:
            if base == normalize_text(a): return canon
    return base

def build_match_index(catalog: list[str]):
    norm_map = {c: normalize_text(c) for c in catalog}
    inv_index = {v: k for k, v in norm_map.items()}
    return norm_map, inv_index

def match_services(src_names: list[str], catalog: list[str], synonyms: dict, score_cutoff=85, scorer=fuzz.WRatio) -> pd.DataFrame:
    _, inv_index = build_match_index(catalog)
    norm_catalog = list(inv_index.keys())
    rows=[]
    for raw in src_names:
        src_norm = apply_synonyms(raw, synonyms)
        cands = process.extract(src_norm, norm_catalog, scorer=scorer, limit=3)
        if cands:
            best_norm, score, _ = cands[0]
            best_canon = inv_index.get(best_norm, best_norm)
            matched = score >= score_cutoff
        else:
            best_canon, score, matched = "", 0, False
        rows.append({
            "source_service": raw,
            "matched_service": best_canon,
            "score": score,
            "is_confident": matched
        })
    return pd.DataFrame(rows)

def read_excel_best_effort(file_bytes: bytes) -> pd.DataFrame:
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    frames=[]
    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=0, dtype=str)
            df = df.dropna(how="all", axis=1).dropna(how="all")
            if not df.empty:
                df["__sheet__"] = sheet
                frames.append(df)
        except: pass
    if not frames: return pd.DataFrame()
    big = pd.concat(frames, ignore_index=True)
    for c in big.columns: big[c] = big[c].astype(str).map(lambda x: x.strip())
    return big


# ========== Model bảng ==========
class PandasModel(QAbstractTableModel):
    def __init__(self, df=pd.DataFrame()):
        super().__init__()
        self._df = df
    def rowCount(self, parent=QModelIndex()): return 0 if self._df is None else len(self._df.index)
    def columnCount(self, parent=QModelIndex()): return 0 if self._df is None else len(self._df.columns)
    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or self._df is None: return None
        if role == Qt.DisplayRole:
            val = self._df.iat[index.row(), index.column()]
            if pd.isna(val): return ""
            if isinstance(val, (pd.Timestamp, datetime.date, datetime.datetime)):
                try: return pd.to_datetime(val).strftime("%d/%m/%Y")
                except Exception: return str(val)
            return str(val)
        return None
    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if self._df is None: return None
        if role == Qt.DisplayRole:
            return str(self._df.columns[section]) if orientation == Qt.Horizontal else str(section + 1)
        return None
    def setDataFrame(self, df: pd.DataFrame):
        self.beginResetModel(); self._df = df; self.endResetModel()

# ========== Workers ==========
class FIVWorker(QThread):
    progress = Signal(int); error = Signal(str); done = Signal(pd.DataFrame, str)
    def __init__(self, eas_path: str, kh_path: str, out_path: str):
        super().__init__(); self.eas_path=eas_path; self.kh_path=kh_path; self.out_path=out_path
    def run(self):
        try:
            self.progress.emit(5)
            with open(self.eas_path, "rb") as f: eas_bytes = f.read()
            df_kh = pd.read_excel(self.kh_path); self.progress.emit(20)
            df_raw = load_and_flatten_eas(eas_bytes); df_eas = clean_eas(df_raw); self.progress.emit(50)
            df_fiv = build_fiv(df_eas, df_kh)
            for c in ['InvoiceDate','DocumentDate','BHS_VATInvocieDate_VATInvoice']:
                df_fiv[c] = pd.to_datetime(df_fiv[c], dayfirst=True, errors='coerce').dt.date
            self.progress.emit(70)
            with pd.ExcelWriter(self.out_path, engine='xlsxwriter', date_format='dd/mm/yyyy') as w:
                df_fiv.to_excel(w, index=False, sheet_name='FIV')
            self.progress.emit(100); self.done.emit(df_fiv, "Đã tạo file FIV thành công.")
        except Exception as e:
            self.error.emit(f"Lỗi FIV: {e}\n{traceback.format_exc()}")

class AgodaWorker(QThread):
    progress = Signal(int); error = Signal(str); done = Signal(pd.DataFrame, str)
    def __init__(self, agoda_path: str, start_date: datetime.date, end_date: datetime.date, out_path: str, chosen_sheet: str|None):
        super().__init__(); self.agoda_path=agoda_path; self.start_date=start_date; self.end_date=end_date; self.out_path=out_path; self.chosen_sheet=chosen_sheet
    def run(self):
        try:
            self.progress.emit(5); required_cols={"Ngày trả phòng","Doanh thu thực","Số tiền bị trừ"}
            xls = pd.ExcelFile(self.agoda_path); self.progress.emit(15)
            candidates=[]; 
            for sh in xls.sheet_names:
                tmp = pd.read_excel(xls, sheet_name=sh, nrows=5)
                if required_cols.issubset(set(tmp.columns)): candidates.append(sh)
            if not candidates: raise ValueError("Không tìm thấy sheet có đủ cột bắt buộc: " + ", ".join(sorted(required_cols)))
            sheet = self.chosen_sheet if self.chosen_sheet in candidates else candidates[0]
            df = pd.read_excel(xls, sheet_name=sheet); self.progress.emit(35)
            df["Ngày trả phòng"] = df["Ngày trả phòng"].apply(parse_vietnamese_date)
            df["Doanh thu thực"] = df["Doanh thu thực"].astype(str).str.replace(",", "", regex=False).str.strip().astype(float)
            df["Số tiền bị trừ"] = df["Số tiền bị trừ"].astype(str).str.replace(",", "", regex=False).str.strip().astype(float)
            start_ts=pd.to_datetime(self.start_date); end_ts=pd.to_datetime(self.end_date)
            mask = (df["Ngày trả phòng"]>=start_ts)&(df["Ngày trả phòng"]<=end_ts)
            df_filtered = df.loc[mask].copy()
            df_filtered = df_filtered[(df_filtered["Doanh thu thực"]>0)&(df_filtered["Số tiền bị trừ"]>0)]
            df_filtered = df_filtered.loc[:, ~df_filtered.columns.str.contains("^Unnamed")]; self.progress.emit(70)
            with pd.ExcelWriter(self.out_path, engine="xlsxwriter") as w: df_filtered.to_excel(w, index=False, sheet_name="Agoda")
            self.progress.emit(100); self.done.emit(df_filtered, f"Đã xuất file Agoda: {os.path.basename(self.out_path)} (sheet: {sheet})")
        except Exception as e:
            self.error.emit(f"Lỗi Agoda: {e}\n{traceback.format_exc()}")

# ========== Tabs ==========
class FIVTab(QWidget):
    def __init__(self):
        super().__init__()
        self.worker: FIVWorker | None = None
        self.df_preview = pd.DataFrame()
        self.selected_out_path: str | None = None  # lưu path người dùng chọn từ "Save as"

        # Widgets
        self.in_eas = QLineEdit(); self.in_eas.setPlaceholderText("Chọn file EAS.xlsx")
        self.btn_eas = QPushButton("Chọn EAS.xlsx")
        self.in_kh = QLineEdit(); self.in_kh.setPlaceholderText("Chọn file KH.xlsx")
        self.btn_kh = QPushButton("Chọn KH.xlsx")

        # KHÔNG còn QLineEdit cho Output; chỉ còn nút Save as
        self.btn_out = QPushButton("Save as")
        self.btn_run = QPushButton("Chạy tạo FIV")
        self.pbar = QProgressBar(); self.pbar.setValue(0)

        # layout
        layout = QVBoxLayout(self)
        box = QGroupBox("FIV Generator")
        v = QVBoxLayout(box)

        # hàng 1
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("EAS.xlsx:"))
        row1.addWidget(self.in_eas)
        row1.addWidget(self.btn_eas)

        # hàng 2
        row2 = QHBoxLayout()
        row2.addWidget(QLabel("KH.xlsx:"))
        row2.addWidget(self.in_kh)
        row2.addWidget(self.btn_kh)

        # hàng 3 (Output chỉ còn nhãn + nút Save as)
        row3 = QHBoxLayout()
        row3.addWidget(QLabel("Output:"))
        row3.addSpacing(10)     
        row3.addWidget(self.btn_out) 

        v.addLayout(row1)
        v.addLayout(row2)
        v.addLayout(row3)
        v.addWidget(self.pbar)
        v.addWidget(self.btn_run)
        layout.addWidget(box)

        # events
        self.btn_eas.clicked.connect(self.pick_eas)
        self.btn_kh.clicked.connect(self.pick_kh)
        self.btn_out.clicked.connect(self.pick_out)
        self.btn_run.clicked.connect(self.run)

    def pick_eas(self):
        path, _ = QFileDialog.getOpenFileName(self, "Chọn file EAS.xlsx", filter="Excel (*.xlsx *.xls)")
        if path:
            self.in_eas.setText(path)

    def pick_kh(self):
        path, _ = QFileDialog.getOpenFileName(self, "Chọn file KH.xlsx", filter="Excel (*.xlsx *.xls)")
        if path:
            self.in_kh.setText(path)

    def pick_out(self):
        # mặc định gợi ý tên Completed_FIV.xlsx trong thư mục Documents của user
        suggested = os.path.join(os.path.expanduser("~"), "Documents", "Completed_FIV.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "Save as", suggested, filter="Excel (*.xlsx)")
        if path:
            # đảm bảo phần mở rộng .xlsx
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self.selected_out_path = path
            QMessageBox.information(self, "Output đã chọn", f"Sẽ lưu tại:\n{self.selected_out_path}")

    def run(self):
        eas = self.in_eas.text().strip()
        kh = self.in_kh.text().strip()
        if not eas or not kh:
            QMessageBox.warning(self, "Thiếu thông tin", "Vui lòng chọn đủ EAS.xlsx và KH.xlsx")
            return
        # Bắt buộc chọn Save as trước khi chạy
        if not self.selected_out_path:
            QMessageBox.warning(self, "Chưa chọn nơi lưu", "Vui lòng bấm 'Save as' để chọn file xuất.")
            return

        self.btn_run.setEnabled(False)
        self.worker = FIVWorker(eas, kh, self.selected_out_path)
        self.worker.progress.connect(self.pbar.setValue)
        self.worker.error.connect(lambda msg: (self.btn_run.setEnabled(True), QMessageBox.critical(self, "Lỗi", msg)))
        self.worker.done.connect(lambda df, m: (self.btn_run.setEnabled(True), QMessageBox.information(self, "Xong", m)))
        self.worker.start()

class AgodaTab(QWidget):
    def __init__(self):
        super().__init__(); self.worker=None
        today = datetime.date.today(); default_start = today - datetime.timedelta(days=7); default_end = today
        self.in_agoda = QLineEdit(placeholderText="Chọn file Agoda.xlsx"); self.btn_agoda = QPushButton("Chọn file Agoda")
        self.dtp_start = QDateEdit(); self.dtp_start.setCalendarPopup(True); self.dtp_start.setDate(default_start)
        self.dtp_end = QDateEdit();   self.dtp_end.setCalendarPopup(True);   self.dtp_end.setDate(default_end)
        self.sheet_select = QComboBox(); self.sheet_select.addItem("(Tự động phát hiện)")
        self.btn_detect_sheets = QPushButton("Quét sheet hợp lệ")
        self.out_path = QLineEdit(placeholderText="Nơi lưu Agoda_processed_YYYYMMDD_YYYYMMDD.xlsx"); self.btn_out = QPushButton("Chọn nơi lưu")
        self.pbar = QProgressBar(); self.btn_run = QPushButton("Chạy xử lý Agoda")

        layout = QVBoxLayout(self); box = QGroupBox("Agoda LCB Processor"); v = QVBoxLayout(box)
        row1 = QHBoxLayout(); row1.addWidget(QLabel("Agoda.xlsx:")); row1.addWidget(self.in_agoda); row1.addWidget(self.btn_agoda)
        row2 = QHBoxLayout(); row2.addWidget(QLabel("Từ ngày:")); row2.addWidget(self.dtp_start); row2.addWidget(QLabel("Đến ngày:")); row2.addWidget(self.dtp_end)
        row3 = QHBoxLayout(); row3.addWidget(QLabel("Sheet:")); row3.addWidget(self.sheet_select); row3.addWidget(self.btn_detect_sheets)
        row4 = QHBoxLayout(); row4.addWidget(QLabel("Output:")); row4.addWidget(self.out_path); row4.addWidget(self.btn_out)
        v.addLayout(row1); v.addLayout(row2); v.addLayout(row3); v.addLayout(row4); v.addWidget(self.pbar); v.addWidget(self.btn_run); layout.addWidget(box)

        self.btn_agoda.clicked.connect(self.pick_agoda); self.btn_out.clicked.connect(self.pick_out)
        self.btn_detect_sheets.clicked.connect(self.detect_sheets); self.btn_run.clicked.connect(self.run)

    def pick_agoda(self):
        p,_ = QFileDialog.getOpenFileName(self, "Chọn file Agoda.xlsx", filter="Excel (*.xlsx *.xls)")
        if p: self.in_agoda.setText(p)
    def pick_out(self):
        p,_ = QFileDialog.getSaveFileName(self, "Lưu Agoda_processed.xlsx", filter="Excel (*.xlsx)")
        if p: self.out_path.setText(p)

    def detect_sheets(self):
        path = self.in_agoda.text().strip()
        if not path: QMessageBox.warning(self, "Thiếu file", "Hãy chọn file Agoda trước."); return
        try:
            xls = pd.ExcelFile(path); required_cols={"Ngày trả phòng","Doanh thu thực","Số tiền bị trừ"}; candidates=[]
            for sh in xls.sheet_names:
                tmp = pd.read_excel(xls, sheet_name=sh, nrows=5)
                if required_cols.issubset(set(tmp.columns)): candidates.append(sh)
            self.sheet_select.clear(); self.sheet_select.addItem("(Tự động phát hiện)")
            if candidates:
                self.sheet_select.addItems(candidates); QMessageBox.information(self, "OK", f"Tìm thấy {len(candidates)} sheet hợp lệ.")
            else:
                QMessageBox.warning(self, "Không thấy", "Không có sheet nào đủ cột bắt buộc.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi đọc file: {e}")

    def run(self):
        agoda = self.in_agoda.text().strip()
        if not agoda: QMessageBox.warning(self, "Thiếu file", "Vui lòng chọn file Agoda.xlsx"); return
        start = self.dtp_start.date().toPython(); end = self.dtp_end.date().toPython()
        if start > end: QMessageBox.warning(self, "Sai ngày", "Ngày bắt đầu phải <= ngày kết thúc."); return
        outp = self.out_path.text().strip() or os.path.join(os.path.expanduser("~"), f"Agoda_processed_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx")
        chosen = self.sheet_select.currentText(); chosen = None if "(Tự động phát hiện)" in chosen else chosen

        self.btn_run.setEnabled(False)
        self.worker = AgodaWorker(agoda, start, end, outp, chosen)
        self.worker.progress.connect(self.pbar.setValue)
        self.worker.error.connect(lambda msg: (self.btn_run.setEnabled(True), QMessageBox.critical(self, "Lỗi", msg)))
        self.worker.done.connect(lambda df, m: (self.btn_run.setEnabled(True), QMessageBox.information(self, "Xong", m)))
        self.worker.start()
class HospitalTab(QWidget):
    def __init__(self):
        super().__init__()
        self.synonyms = DEFAULT_SYNONYMS.copy()
        self.catalog: list[str] = []
        self.template_path = None
        self.hosp_files = []  # [(label, path, df, cfg)]
        self.cutoff = 85
        self.scorer = "WRatio"
        self.final_df = pd.DataFrame()

        # ==== UI ====
        main = QVBoxLayout(self)
        box = QGroupBox("Báo giá bệnh viện → Ghép vào Template cố định")
        v = QVBoxLayout(box)

        # Row: Template + 3 bệnh viện
        r1 = QHBoxLayout()
        self.txt_tpl = QLineEdit(placeholderText="Chọn Template.xlsx (có thể có sheet CATALOG/service_name)")
        self.btn_tpl = QPushButton("Chọn Template")
        r1.addWidget(QLabel("Template:")); r1.addWidget(self.txt_tpl); r1.addWidget(self.btn_tpl)

        r2 = QHBoxLayout()
        self.txt_h1 = QLineEdit(placeholderText="Chọn BV A.xlsx"); self.btn_h1 = QPushButton("Chọn BV A")
        self.txt_h2 = QLineEdit(placeholderText="Chọn BV B.xlsx"); self.btn_h2 = QPushButton("Chọn BV B")
        self.txt_h3 = QLineEdit(placeholderText="Chọn BV C.xlsx"); self.btn_h3 = QPushButton("Chọn BV C")
        r2.addWidget(self.txt_h1); r2.addWidget(self.btn_h1)
        r2.addWidget(self.txt_h2); r2.addWidget(self.btn_h2)
        r2.addWidget(self.txt_h3); r2.addWidget(self.btn_h3)

        # Row: Synonyms & cutoff
        r3 = QHBoxLayout()
        self.txt_syn = QLineEdit()
        self.txt_syn.setPlaceholderText("Từ điển đồng nghĩa (format: 'chuẩn: biến thể1, biến thể2; ...'). Để trống = mặc định.")
        self.cmb_scorer = QComboBox(); self.cmb_scorer.addItems(["WRatio","TokenSortRatio","TokenSetRatio"])
        self.cmb_cutoff = QComboBox(); self.cmb_cutoff.addItems([str(x) for x in range(60,101)])
        self.cmb_cutoff.setCurrentText(str(self.cutoff))
        r3.addWidget(QLabel("Synonyms:")); r3.addWidget(self.txt_syn, 2)
        r3.addWidget(QLabel("Scorer:")); r3.addWidget(self.cmb_scorer)
        r3.addWidget(QLabel("Cutoff:")); r3.addWidget(self.cmb_cutoff)

        # Buttons
        r4 = QHBoxLayout()
        self.btn_detect = QPushButton("Đọc Template & Chuẩn bị danh mục")
        self.btn_config_cols = QPushButton("Chọn cột cho từng BV")
        self.btn_run = QPushButton("Ghép & Xem kết quả")
        self.btn_save_tpl = QPushButton("Lưu Excel (thêm sheet TONG_HOP)")
        r4.addWidget(self.btn_detect); r4.addWidget(self.btn_config_cols); r4.addWidget(self.btn_run); r4.addWidget(self.btn_save_tpl)

        # Preview
        self.tbl_preview = QTabWidget()
        self.tbl_result = PandasModel()
        self.tbl_view = QLabel("Chưa có dữ liệu. Hãy nhấn 'Ghép & Xem kết quả'.")
        self.tbl_view.setAlignment(Qt.AlignCenter)

        v.addLayout(r1); v.addLayout(r2); v.addLayout(r3); v.addLayout(r4)
        main.addWidget(box)
        main.addWidget(self.tbl_preview)
        main.addWidget(self.tbl_view)

        # events
        self.btn_tpl.clicked.connect(lambda: self.pick_file(self.txt_tpl))
        self.btn_h1.clicked.connect(lambda: self.pick_file(self.txt_h1))
        self.btn_h2.clicked.connect(lambda: self.pick_file(self.txt_h2))
        self.btn_h3.clicked.connect(lambda: self.pick_file(self.txt_h3))
        self.btn_detect.clicked.connect(self.detect_template_catalog)
        self.btn_config_cols.clicked.connect(self.configure_columns)
        self.btn_run.clicked.connect(self.run_match)
        self.btn_save_tpl.clicked.connect(self.save_into_template)

    def pick_file(self, txt: QLineEdit):
        p, _ = QFileDialog.getOpenFileName(self, "Chọn file Excel", filter="Excel (*.xlsx *.xls)")
        if p: txt.setText(p)

    def parse_synonyms(self):
        raw = self.txt_syn.text().strip()
        if not raw: 
            self.synonyms = DEFAULT_SYNONYMS.copy(); return
        syn = {}
        # ví dụ: "kham tong quat: kham SK, kham tong; x quang: xquang"
        for segment in re.split(r"[;\\n]+", raw):
            if ":" in segment:
                left, right = segment.split(":", 1)
                canon = normalize_text(left)
                alts = [normalize_text(x) for x in right.split(",") if x.strip()]
                if canon: syn[canon] = alts
        self.synonyms = syn if syn else DEFAULT_SYNONYMS.copy()

    def detect_template_catalog(self):
        self.parse_synonyms()
        self.catalog = []
        tpl = self.txt_tpl.text().strip()
        if not tpl:
            QMessageBox.warning(self, "Thiếu Template", "Hãy chọn Template.xlsx"); return
        self.template_path = tpl
        try:
            xls = pd.ExcelFile(tpl)
            cat_sheet = None
            for s in xls.sheet_names:
                if normalize_text(s) in {"catalog","danh muc","danhmuc","dm"}:
                    cat_sheet = s; break
            if cat_sheet:
                df = pd.read_excel(xls, sheet_name=cat_sheet, dtype=str)
                col = None
                for c in df.columns:
                    if normalize_text(c) in {"service_name","service","ten dich vu","dich vu","noi dung"}:
                        col = c; break
                if col:
                    self.catalog = (df[col].dropna().astype(str).tolist())
            if not self.catalog:
                # không có catalog sẵn → hỏi nhập nhanh
                text, ok = QInputDialog.getMultiLineText(self, "Danh mục chuẩn", "Mỗi dòng 1 dịch vụ:")
                if ok:
                    self.catalog = [x.strip() for x in text.splitlines() if x.strip()]
            if not self.catalog:
                QMessageBox.warning(self, "Chưa có danh mục", "Không có danh mục dịch vụ chuẩn."); return
            QMessageBox.information(self, "OK", f"Đã nạp {len(self.catalog)} dịch vụ chuẩn.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi đọc Template: {e}")

    def configure_columns(self):
        # Đọc 3 file BV, đoán cột và cho xem preview
        self.hosp_files.clear()
        self.tbl_preview.clear()
        for txt, label in [(self.txt_h1,"BV_A"),(self.txt_h2,"BV_B"),(self.txt_h3,"BV_C")]:
            p = txt.text().strip()
            if not p: continue
            try:
                with open(p, "rb") as f: b = f.read()
                df = read_excel_best_effort(b)
                if df.empty: 
                    QMessageBox.warning(self, "Cảnh báo", f"{label}: không có dữ liệu hợp lệ."); 
                    continue
                # đoán cột
                guess = {"service":None,"price":None,"unit":None,"dept":None,"note":None,"vat":None}
                for c in df.columns:
                    t = guess_colname(c)
                    if t and (guess[t] is None):
                        guess[t] = c
                # popup lựa chọn cột tối thiểu
                svc, ok1 = QInputDialog.getItem(self, f"{label} - Cột TÊN DV", "Chọn cột Tên DV:", [c for c in df.columns], 
                                                editable=False, current= (list(df.columns).index(guess["service"]) if guess["service"] in df.columns else 0))
                if not ok1: continue
                prc, ok2 = QInputDialog.getItem(self, f"{label} - Cột GIÁ", "Chọn cột Giá:", [c for c in df.columns], 
                                                editable=False, current= (list(df.columns).index(guess["price"]) if guess["price"] in df.columns else 0))
                if not ok2: continue
                cfg={"service":svc,"price":prc,"unit":guess.get("unit"),"dept":guess.get("dept"),"note":guess.get("note"),"vat":guess.get("vat")}
                self.hosp_files.append((label, p, df, cfg))
                # preview tab
                prev = df.head(100).copy()
                view = QTableView(); model = PandasModel(prev); view.setModel(model)
                self.tbl_preview.addTab(view, label)
            except Exception as e:
                QMessageBox.critical(self, "Lỗi", f"{label}: {e}")

    def run_match(self):
        if not self.catalog:
            QMessageBox.warning(self, "Thiếu danh mục", "Hãy bấm 'Đọc Template & Chuẩn bị danh mục' trước."); return
        if not self.hosp_files:
            QMessageBox.warning(self, "Thiếu dữ liệu", "Hãy bấm 'Chọn cột cho từng BV' để nạp file."); return
        self.cutoff = int(self.cmb_cutoff.currentText())
        self.scorer = self.cmb_scorer.currentText()
        scorer = {"WRatio": fuzz.WRatio, "TokenSortRatio": fuzz.token_sort_ratio, "TokenSetRatio": fuzz.token_set_ratio}[self.scorer]

        all_rows=[]
        for label, path, df, cfg in self.hosp_files:
            tmp = df[[cfg["service"], cfg["price"]]].copy()
            tmp = tmp.rename(columns={cfg["service"]:"source_service", cfg["price"]:"source_price"})
            for k in ["unit","dept","note","vat"]:
                tmp[k] = df[cfg[k]] if cfg.get(k) and cfg[k] in df.columns else None
            # parse price
            def parse_price(x):
                if pd.isna(x): return None
                s = str(x).replace(" ","").replace(",","")
                s = re.sub(r"[^0-9.]+","", s)
                try: return float(s) if s!="" else None
                except: return None
            tmp["price_numeric"] = tmp["source_price"].map(parse_price)
            tmp = tmp[tmp["source_service"].astype(str).str.strip()!=""]

            mdf = match_services(tmp["source_service"].astype(str).tolist(), self.catalog, self.synonyms, self.cutoff, scorer)
            merged = pd.concat([tmp.reset_index(drop=True), mdf[["matched_service","score","is_confident"]]], axis=1)
            merged["hospital"] = label
            all_rows.append(merged)

        result_df = pd.concat(all_rows, ignore_index=True)
        # lấy giá tốt nhất theo score cho mỗi (hospital, matched_service)
        agg = (result_df.dropna(subset=["matched_service"])
               .sort_values(["hospital","matched_service","score"], ascending=[True,True,False])
               .groupby(["hospital","matched_service"], as_index=False).first())

        # pivot sang dạng rộng
        target = pd.DataFrame({"service_name": self.catalog})
        wide=[]
        for hosp in agg["hospital"].unique():
            sub = agg[agg["hospital"]==hosp][["matched_service","price_numeric"]].rename(columns={"price_numeric":f"{hosp}_price"})
            wide.append(sub)
        for wf in wide:
            target = target.merge(wf, left_on="service_name", right_on="matched_service", how="left").drop(columns=["matched_service"])

        # unit/vat nếu đồng nhất
        for meta in ["unit","vat"]:
            picks = agg[["matched_service", meta]].dropna().groupby("matched_service")[meta].nunique()
            single = picks[picks==1].index.tolist()
            meta_map = (agg[agg["matched_service"].isin(single)][["matched_service",meta]]
                        .dropna().drop_duplicates("matched_service").set_index("matched_service")[meta].to_dict())
            target[meta] = target["service_name"].map(meta_map)

        price_cols = [c for c in target.columns if c.endswith("_price")]
        cols = ["service_name","unit","vat"] + price_cols
        self.final_df = target[cols]

        # hiển thị
        tv = QTableView(); tv.setModel(PandasModel(self.final_df))
        self.tbl_preview.addTab(tv, "Kết quả")
        self.tbl_view.setText("")

        # cảnh báo match thấp
        low = (result_df["is_confident"]==False).sum()
        if low>0:
            QMessageBox.warning(self, "Lưu ý", f"Có {low} dòng match dưới ngưỡng {self.cutoff}. Hãy xem tab 'Kết quả'.")

    def save_into_template(self):
        if self.final_df.empty:
            QMessageBox.warning(self, "Chưa có dữ liệu", "Hãy chạy 'Ghép & Xem kết quả' trước."); return
        if not self.template_path:
            QMessageBox.warning(self, "Thiếu Template", "Chưa chọn Template để chèn sheet."); return
        out, _ = QFileDialog.getSaveFileName(self, "Lưu file kết quả", filter="Excel (*.xlsx)")
        if not out: return
        if not out.lower().endswith(".xlsx"): out += ".xlsx"

        try:
            from openpyxl import load_workbook
            with open(self.template_path, "rb") as f: tpl_bytes = f.read()
            # copy template sang output
            wb = load_workbook(io.BytesIO(tpl_bytes))
            if "TONG_HOP" in wb.sheetnames:
                ws = wb["TONG_HOP"]; wb.remove(ws)
            wb.create_sheet("TONG_HOP"); 
            bio = io.BytesIO(); wb.save(bio); bio.seek(0)

            with pd.ExcelWriter(out, engine="openpyxl", mode="w") as w:
                # ghi lại toàn bộ workbook (đã có sheet trống TONG_HOP)
                wb2 = load_workbook(bio)
                wb2.save(out)
            with pd.ExcelWriter(out, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w2:
                self.final_df.to_excel(w2, index=False, sheet_name="TONG_HOP")
            QMessageBox.information(self, "OK", f"Đã lưu {os.path.basename(out)} (thêm sheet TONG_HOP).")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể lưu: {e}")

# ========== Main Window ==========
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon(resource_path("app.ico")))
        self.setWindowTitle("Automation Tools (Desktop)")
        self.setMinimumWidth(760); self.setMinimumHeight(420)
        # đặt icon (đặt file app.ico cạnh app_gui.py)
        if os.path.exists("app.ico"): self.setWindowIcon(QIcon("app.ico"))

        tabs = QTabWidget(); tabs.addTab(FIVTab(), "Senspa: FIV Generator"); tabs.addTab(AgodaTab(), "Agoda LCB")
        self.setCentralWidget(tabs), tabs.addTab(HospitalTab(), "Báo giá BV -> Template") 

        menubar = self.menuBar(); fileMenu = menubar.addMenu("File")
        act_about = QAction("About", self); act_about.triggered.connect(self.show_about); fileMenu.addAction(act_about)

    def show_about(self):
        QMessageBox.information(self, "About", "Automation Tools (Desktop)\nUI PySide6 với theme màu & icon.\n© Your Company")

def main():
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    app.setStyleSheet(APP_STYLES)
    w = MainWindow(); w.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
